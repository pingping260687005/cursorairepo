"""
Excel写入工具类

提供写入Excel文件的各种功能，包括：
- 写入DataFrame到工作表
- 写入数据到指定范围
- 创建新的工作表
- 设置单元格格式
- 保存Excel文件
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Union, List, Dict, Optional, Any
import os


class ExcelWriter:
    """Excel文件写入工具类"""
    
    def __init__(self, file_path: str, create_if_not_exists: bool = True):
        """
        初始化ExcelWriter
        
        Args:
            file_path (str): Excel或CSV文件路径
            create_if_not_exists (bool): 如果文件不存在是否创建，默认为True
        """
        self.file_path = file_path
        self.workbook = None
        self.is_csv = file_path.endswith('.csv')
        
        if not self.is_csv:
            self._load_or_create_workbook(create_if_not_exists)
    
    def _load_or_create_workbook(self, create_if_not_exists: bool):
        """加载或创建工作簿"""
        try:
            if os.path.exists(self.file_path):
                self.workbook = openpyxl.load_workbook(self.file_path)
            elif create_if_not_exists:
                self.workbook = openpyxl.Workbook()
                # 删除默认的工作表
                if 'Sheet' in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook['Sheet'])
            else:
                raise FileNotFoundError(f"文件不存在: {self.file_path}")
        except Exception as e:
            raise Exception(f"无法加载或创建Excel文件: {e}")
    
    def write_dataframe(self, df: pd.DataFrame, 
                       sheet_name: str = 'Sheet1',
                       start_row: int = 1,
                       start_col: int = 1,
                       include_index: bool = False,
                       include_header: bool = True,
                       encoding: str = 'utf-8',
                       delimiter: str = ',') -> None:
        """
        将DataFrame写入工作表
        
        Args:
            df (pd.DataFrame): 要写入的数据
            sheet_name (str): 工作表名称
            start_row (int): 起始行号，默认为1
            start_col (int): 起始列号，默认为1
            include_index (bool): 是否包含索引，默认为False
            include_header (bool): 是否包含表头，默认为True
            encoding (str): CSV文件编码，默认为utf-8
            delimiter (str): CSV分隔符，默认为逗号
        """
        try:
            if self.is_csv:
                # CSV文件写入
                df.to_csv(
                    self.file_path,
                    index=include_index,
                    header=include_header,
                    encoding=encoding,
                    sep=delimiter
                )
            else:
                # 如果工作表不存在，创建它
                if sheet_name not in self.workbook.sheetnames:
                    self.workbook.create_sheet(sheet_name)
                
                worksheet = self.workbook[sheet_name]
                
                # 将DataFrame转换为行数据
                for r in dataframe_to_rows(df, index=include_index, header=include_header):
                    for c_idx, value in enumerate(r):
                        cell = worksheet.cell(row=start_row, column=start_col + c_idx)
                        cell.value = value
                    start_row += 1
                
        except Exception as e:
            raise Exception(f"写入DataFrame失败: {e}")
    
    def write_data_to_range(self, data: List[List[Any]], 
                           sheet_name: str,
                           start_cell: str) -> None:
        """
        将数据写入指定范围
        
        Args:
            data (List[List[Any]]): 二维数据列表
            sheet_name (str): 工作表名称
            start_cell (str): 起始单元格，如'A1'
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet_name)
            
            worksheet = self.workbook[sheet_name]
            
            # 解析起始单元格
            start_col = openpyxl.utils.column_index_from_string(start_cell[0])
            start_row = int(start_cell[1:])
            
            # 写入数据
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    cell = worksheet.cell(row=start_row + row_idx, column=start_col + col_idx)
                    cell.value = value
                    
        except Exception as e:
            raise Exception(f"写入指定范围数据失败: {e}")
    
    def write_cell_value(self, sheet_name: str, cell_address: str, value: Any) -> None:
        """
        写入指定单元格的值
        
        Args:
            sheet_name (str): 工作表名称
            cell_address (str): 单元格地址，如'A1'
            value (Any): 要写入的值
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet_name)
            
            worksheet = self.workbook[sheet_name]
            worksheet[cell_address] = value
            
        except Exception as e:
            raise Exception(f"写入单元格值失败: {e}")
    
    def write_column(self, sheet_name: str, column: str, 
                    data: List[Any], start_row: int = 1) -> None:
        """
        写入指定列的数据
        
        Args:
            sheet_name (str): 工作表名称
            column (str): 列名，如'A'或'B'
            data (List[Any]): 要写入的数据列表
            start_row (int): 起始行号，默认为1
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet_name)
            
            worksheet = self.workbook[sheet_name]
            
            for idx, value in enumerate(data):
                cell = worksheet[f"{column}{start_row + idx}"]
                cell.value = value
                
        except Exception as e:
            raise Exception(f"写入列数据失败: {e}")
    
    def write_row(self, sheet_name: str, row: int, 
                 data: List[Any], start_col: int = 1) -> None:
        """
        写入指定行的数据
        
        Args:
            sheet_name (str): 工作表名称
            row (int): 行号
            data (List[Any]): 要写入的数据列表
            start_col (int): 起始列号，默认为1
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet_name)
            
            worksheet = self.workbook[sheet_name]
            
            for idx, value in enumerate(data):
                cell = worksheet.cell(row=row, column=start_col + idx)
                cell.value = value
                
        except Exception as e:
            raise Exception(f"写入行数据失败: {e}")
    
    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> None:
        """
        创建新的工作表
        
        Args:
            sheet_name (str): 工作表名称
            index (int, optional): 插入位置，默认为最后
        """
        try:
            if sheet_name in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 已存在")
            
            self.workbook.create_sheet(sheet_name, index)
            
        except Exception as e:
            raise Exception(f"创建工作表失败: {e}")
    
    def delete_sheet(self, sheet_name: str) -> None:
        """
        删除工作表
        
        Args:
            sheet_name (str): 工作表名称
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
            if len(self.workbook.sheetnames) <= 1:
                raise ValueError("不能删除最后一个工作表")
            
            self.workbook.remove(self.workbook[sheet_name])
            
        except Exception as e:
            raise Exception(f"删除工作表失败: {e}")
    
    def set_cell_format(self, sheet_name: str, cell_address: str,
                       font: Optional[Font] = None,
                       fill: Optional[PatternFill] = None,
                       border: Optional[Border] = None,
                       alignment: Optional[Alignment] = None) -> None:
        """
        设置单元格格式
        
        Args:
            sheet_name (str): 工作表名称
            cell_address (str): 单元格地址
            font (Font, optional): 字体格式
            fill (PatternFill, optional): 填充格式
            border (Border, optional): 边框格式
            alignment (Alignment, optional): 对齐格式
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
            worksheet = self.workbook[sheet_name]
            cell = worksheet[cell_address]
            
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
                
        except Exception as e:
            raise Exception(f"设置单元格格式失败: {e}")
    
    def set_column_width(self, sheet_name: str, column: str, width: float) -> None:
        """
        设置列宽
        
        Args:
            sheet_name (str): 工作表名称
            column (str): 列名，如'A'或'B'
            width (float): 列宽
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
            worksheet = self.workbook[sheet_name]
            worksheet.column_dimensions[column].width = width
            
        except Exception as e:
            raise Exception(f"设置列宽失败: {e}")
    
    def set_row_height(self, sheet_name: str, row: int, height: float) -> None:
        """
        设置行高
        
        Args:
            sheet_name (str): 工作表名称
            row (int): 行号
            height (float): 行高
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
            worksheet = self.workbook[sheet_name]
            worksheet.row_dimensions[row].height = height
            
        except Exception as e:
            raise Exception(f"设置行高失败: {e}")
    
    def freeze_panes(self, sheet_name: str, cell_address: str) -> None:
        """
        冻结窗格
        
        Args:
            sheet_name (str): 工作表名称
            cell_address (str): 冻结位置，如'B2'
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
            worksheet = self.workbook[sheet_name]
            worksheet.freeze_panes = cell_address
            
        except Exception as e:
            raise Exception(f"冻结窗格失败: {e}")
    
    def save(self, file_path: Optional[str] = None) -> None:
        """
        保存Excel或CSV文件
        
        Args:
            file_path (str, optional): 保存路径，默认为初始化时的路径
        """
        try:
            save_path = file_path or self.file_path
            
            # 确保目录存在
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            if self.is_csv:
                # CSV文件已经在write_dataframe中保存了
                pass
            else:
                self.workbook.save(save_path)
            
        except Exception as e:
            raise Exception(f"保存文件失败: {e}")
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()

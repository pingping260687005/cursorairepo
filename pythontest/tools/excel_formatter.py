"""
Excel格式化工具类

提供格式化Excel文件的各种功能，包括：
- 设置字体样式
- 设置单元格颜色和背景
- 设置边框样式
- 设置对齐方式
- 设置数字格式
- 条件格式化
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule, CellIsRule
from openpyxl.utils import get_column_letter
from typing import Union, List, Dict, Optional, Any, Tuple
import os


class ExcelFormatter:
    """Excel格式化工具类"""
    
    def __init__(self, file_path: str):
        """
        初始化ExcelFormatter
        
        Args:
            file_path (str): Excel文件路径
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
    
    def set_font(self, sheet_name: str, 
                cell_range: str,
                font_name: str = 'Arial',
                font_size: int = 11,
                bold: bool = False,
                italic: bool = False,
                underline: bool = False,
                color: str = '000000') -> None:
        """
        设置字体样式
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围，如'A1:C10'
            font_name (str): 字体名称，默认为Arial
            font_size (int): 字体大小，默认为11
            bold (bool): 是否粗体，默认为False
            italic (bool): 是否斜体，默认为False
            underline (bool): 是否下划线，默认为False
            color (str): 字体颜色，默认为黑色
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            font = Font(
                name=font_name,
                size=font_size,
                bold=bold,
                italic=italic,
                underline='single' if underline else None,
                color=color
            )
            
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.font = font
                    
        except Exception as e:
            raise Exception(f"设置字体样式失败: {e}")
    
    def set_fill_color(self, sheet_name: str, 
                      cell_range: str,
                      fill_color: str = 'FFFFFF',
                      pattern_type: str = 'solid') -> None:
        """
        设置填充颜色
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围
            fill_color (str): 填充颜色，默认为白色
            pattern_type (str): 图案类型，默认为solid
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type=pattern_type
            )
            
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.fill = fill
                    
        except Exception as e:
            raise Exception(f"设置填充颜色失败: {e}")
    
    def set_border(self, sheet_name: str, 
                  cell_range: str,
                  border_style: str = 'thin',
                  border_color: str = '000000') -> None:
        """
        设置边框
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围
            border_style (str): 边框样式，默认为thin
            border_color (str): 边框颜色，默认为黑色
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            side = Side(border_style=border_style, color=border_color)
            border = Border(left=side, right=side, top=side, bottom=side)
            
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.border = border
                    
        except Exception as e:
            raise Exception(f"设置边框失败: {e}")
    
    def set_alignment(self, sheet_name: str, 
                     cell_range: str,
                     horizontal: str = 'left',
                     vertical: str = 'bottom',
                     wrap_text: bool = False) -> None:
        """
        设置对齐方式
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围
            horizontal (str): 水平对齐，'left', 'center', 'right'
            vertical (str): 垂直对齐，'top', 'center', 'bottom'
            wrap_text (bool): 是否自动换行，默认为False
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                wrap_text=wrap_text
            )
            
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.alignment = alignment
                    
        except Exception as e:
            raise Exception(f"设置对齐方式失败: {e}")
    
    def set_number_format(self, sheet_name: str, 
                         cell_range: str,
                         number_format: str) -> None:
        """
        设置数字格式
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围
            number_format (str): 数字格式，如'#,##0.00', '0%', 'yyyy-mm-dd'
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.number_format = number_format
                    
        except Exception as e:
            raise Exception(f"设置数字格式失败: {e}")
    
    def format_header(self, sheet_name: str, 
                     header_range: str = 'A1:Z1',
                     font_size: int = 12,
                     bold: bool = True,
                     fill_color: str = 'D3D3D3',
                     text_color: str = '000000') -> None:
        """
        格式化表头
        
        Args:
            sheet_name (str): 工作表名称
            header_range (str): 表头范围，默认为A1:Z1
            font_size (int): 字体大小，默认为12
            bold (bool): 是否粗体，默认为True
            fill_color (str): 背景颜色，默认为浅灰色
            text_color (str): 文字颜色，默认为黑色
        """
        try:
            # 设置字体
            self.set_font(sheet_name, header_range, 
                         font_size=font_size, bold=bold, color=text_color)
            
            # 设置背景色
            self.set_fill_color(sheet_name, header_range, fill_color)
            
            # 设置居中对齐
            self.set_alignment(sheet_name, header_range, 
                             horizontal='center', vertical='center')
            
            # 设置边框
            self.set_border(sheet_name, header_range)
            
        except Exception as e:
            raise Exception(f"格式化表头失败: {e}")
    
    def auto_adjust_column_width(self, sheet_name: str, 
                               columns: Optional[List[str]] = None) -> None:
        """
        自动调整列宽
        
        Args:
            sheet_name (str): 工作表名称
            columns (List[str], optional): 要调整的列，默认为所有列
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            if columns is None:
                # 获取所有有数据的列
                max_col = worksheet.max_column
                columns = [get_column_letter(i) for i in range(1, max_col + 1)]
            
            for column in columns:
                max_length = 0
                
                # 计算列中最大字符长度
                for cell in worksheet[column]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # 设置列宽，最小10，最大50
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column].width = adjusted_width
                
        except Exception as e:
            raise Exception(f"自动调整列宽失败: {e}")
    
    def auto_adjust_row_height(self, sheet_name: str, 
                             rows: Optional[List[int]] = None) -> None:
        """
        自动调整行高
        
        Args:
            sheet_name (str): 工作表名称
            rows (List[int], optional): 要调整的行，默认为所有行
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            if rows is None:
                max_row = worksheet.max_row
                rows = list(range(1, max_row + 1))
            
            for row in rows:
                max_height = 0
                
                # 计算行中最大字符高度
                for cell in worksheet[row]:
                    if cell.value:
                        # 估算行高（每行约15像素）
                        lines = str(cell.value).count('\n') + 1
                        max_height = max(max_height, lines * 15)
                
                # 设置行高，最小15，最大100
                adjusted_height = min(max(max_height, 15), 100)
                worksheet.row_dimensions[row].height = adjusted_height
                
        except Exception as e:
            raise Exception(f"自动调整行高失败: {e}")
    
    def apply_conditional_formatting(self, sheet_name: str, 
                                   cell_range: str,
                                   rule_type: str,
                                   **kwargs) -> None:
        """
        应用条件格式化
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围
            rule_type (str): 规则类型，'color_scale', 'data_bar', 'icon_set', 'cell_is'
            **kwargs: 其他参数
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            if rule_type == 'color_scale':
                # 颜色刻度
                rule = ColorScaleRule(
                    start_type=kwargs.get('start_type', 'min'),
                    start_color=kwargs.get('start_color', 'FF0000'),
                    mid_type=kwargs.get('mid_type', 'percentile'),
                    mid_value=kwargs.get('mid_value', 50),
                    mid_color=kwargs.get('mid_color', 'FFFF00'),
                    end_type=kwargs.get('end_type', 'max'),
                    end_color=kwargs.get('end_color', '00FF00')
                )
            elif rule_type == 'data_bar':
                # 数据条
                rule = DataBarRule(
                    start_type=kwargs.get('start_type', 'min'),
                    start_value=kwargs.get('start_value', 0),
                    end_type=kwargs.get('end_type', 'max'),
                    end_value=kwargs.get('end_value', 0),
                    color=kwargs.get('color', '0000FF')
                )
            elif rule_type == 'icon_set':
                # 图标集
                rule = IconSetRule(
                    icon_style=kwargs.get('icon_style', '3TrafficLights1'),
                    type=kwargs.get('type', 'percent'),
                    values=kwargs.get('values', [33, 67])
                )
            elif rule_type == 'cell_is':
                # 单元格值条件
                rule = CellIsRule(
                    operator=kwargs.get('operator', 'greaterThan'),
                    formula=kwargs.get('formula', ['0']),
                    fill=kwargs.get('fill', PatternFill(start_color='FFFF00')),
                    font=kwargs.get('font', Font(color='FF0000'))
                )
            else:
                raise ValueError(f"不支持的规则类型: {rule_type}")
            
            worksheet.conditional_formatting.add(cell_range, rule)
            
        except Exception as e:
            raise Exception(f"应用条件格式化失败: {e}")
    
    def create_named_style(self, style_name: str, 
                          font: Optional[Font] = None,
                          fill: Optional[PatternFill] = None,
                          border: Optional[Border] = None,
                          alignment: Optional[Alignment] = None) -> None:
        """
        创建命名样式
        
        Args:
            style_name (str): 样式名称
            font (Font, optional): 字体样式
            fill (PatternFill, optional): 填充样式
            border (Border, optional): 边框样式
            alignment (Alignment, optional): 对齐样式
        """
        try:
            style = NamedStyle(name=style_name)
            
            if font:
                style.font = font
            if fill:
                style.fill = fill
            if border:
                style.border = border
            if alignment:
                style.alignment = alignment
            
            self.workbook.add_named_style(style)
            
        except Exception as e:
            raise Exception(f"创建命名样式失败: {e}")
    
    def apply_named_style(self, sheet_name: str, 
                         cell_range: str,
                         style_name: str) -> None:
        """
        应用命名样式
        
        Args:
            sheet_name (str): 工作表名称
            cell_range (str): 单元格范围
            style_name (str): 样式名称
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.style = style_name
                    
        except Exception as e:
            raise Exception(f"应用命名样式失败: {e}")
    
    def format_as_table(self, sheet_name: str, 
                       table_range: str,
                       table_style: str = 'TableStyleMedium2') -> None:
        """
        格式化为表格样式
        
        Args:
            sheet_name (str): 工作表名称
            table_range (str): 表格范围
            table_style (str): 表格样式，默认为TableStyleMedium2
        """
        try:
            worksheet = self.workbook[sheet_name]
            
            # 创建表格
            table = openpyxl.worksheet.table.Table(
                ref=table_range,
                displayName=f"Table_{sheet_name}",
                tableStyleInfo=openpyxl.worksheet.table.TableStyleInfo(
                    name=table_style,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
            )
            
            worksheet.add_table(table)
            
        except Exception as e:
            raise Exception(f"格式化为表格失败: {e}")
    
    def save(self, file_path: Optional[str] = None) -> None:
        """
        保存Excel文件
        
        Args:
            file_path (str, optional): 保存路径，默认为原文件路径
        """
        try:
            save_path = file_path or self.file_path
            self.workbook.save(save_path)
            
        except Exception as e:
            raise Exception(f"保存Excel文件失败: {e}")
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()

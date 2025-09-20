"""
统一文件格式检测和处理器

提供统一的文件格式检测、读取和写入功能，支持Excel和CSV格式。
"""

import pandas as pd
import openpyxl
import os
from typing import Union, List, Dict, Optional, Any
from enum import Enum


class FileFormat(Enum):
    """支持的文件格式"""
    EXCEL = "excel"
    CSV = "csv"
    UNKNOWN = "unknown"


class FileHandler:
    """统一文件格式处理器"""
    
    def __init__(self, file_path: str):
        """
        初始化FileHandler
        
        Args:
            file_path (str): 文件路径
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        self.file_path = file_path
        self.format = self._detect_format()
        self.workbook = None
        
        if self.format == FileFormat.EXCEL:
            self._load_workbook()
    
    def _detect_format(self) -> FileFormat:
        """
        检测文件格式
        
        Returns:
            FileFormat: 文件格式
        """
        file_extension = os.path.splitext(self.file_path)[1].lower()
        
        if file_extension in ['.xlsx', '.xls']:
            return FileFormat.EXCEL
        elif file_extension == '.csv':
            return FileFormat.CSV
        else:
            return FileFormat.UNKNOWN
    
    def _load_workbook(self):
        """加载Excel工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise Exception(f"无法加载Excel文件: {e}")
    
    def get_format(self) -> FileFormat:
        """
        获取文件格式
        
        Returns:
            FileFormat: 文件格式
        """
        return self.format
    
    def is_excel(self) -> bool:
        """
        是否为Excel文件
        
        Returns:
            bool: 是否为Excel文件
        """
        return self.format == FileFormat.EXCEL
    
    def is_csv(self) -> bool:
        """
        是否为CSV文件
        
        Returns:
            bool: 是否为CSV文件
        """
        return self.format == FileFormat.CSV
    
    def read_data(self, 
                  sheet_name: Optional[str] = None,
                  header_row: int = 0,
                  use_cols: Optional[Union[str, List[int]]] = None,
                  encoding: str = 'utf-8',
                  delimiter: str = ',') -> pd.DataFrame:
        """
        读取数据
        
        Args:
            sheet_name (str, optional): 工作表名称（仅Excel）
            header_row (int): 表头行号，默认为0
            use_cols (str or List[int], optional): 要读取的列，默认为所有列
            encoding (str): CSV文件编码，默认为utf-8
            delimiter (str): CSV分隔符，默认为逗号
            
        Returns:
            pd.DataFrame: 读取的数据
        """
        try:
            if self.format == FileFormat.CSV:
                return pd.read_csv(
                    self.file_path,
                    header=header_row,
                    usecols=use_cols,
                    encoding=encoding,
                    sep=delimiter
                )
            elif self.format == FileFormat.EXCEL:
                if sheet_name is None:
                    sheet_name = self.workbook.sheetnames[0]
                
                return pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    header=header_row,
                    usecols=use_cols
                )
            else:
                raise Exception(f"不支持的文件格式: {self.format}")
                
        except Exception as e:
            raise Exception(f"读取数据失败: {e}")
    
    def write_data(self, 
                   df: pd.DataFrame,
                   sheet_name: str = 'Sheet1',
                   include_index: bool = False,
                   include_header: bool = True,
                   encoding: str = 'utf-8',
                   delimiter: str = ',') -> None:
        """
        写入数据
        
        Args:
            df (pd.DataFrame): 要写入的数据
            sheet_name (str): 工作表名称（仅Excel）
            include_index (bool): 是否包含索引，默认为False
            include_header (bool): 是否包含表头，默认为True
            encoding (str): CSV文件编码，默认为utf-8
            delimiter (str): CSV分隔符，默认为逗号
        """
        try:
            if self.format == FileFormat.CSV:
                df.to_csv(
                    self.file_path,
                    index=include_index,
                    header=include_header,
                    encoding=encoding,
                    sep=delimiter
                )
            elif self.format == FileFormat.EXCEL:
                with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=include_index,
                        header=include_header
                    )
            else:
                raise Exception(f"不支持的文件格式: {self.format}")
                
        except Exception as e:
            raise Exception(f"写入数据失败: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """
        获取工作表名称列表
        
        Returns:
            List[str]: 工作表名称列表
        """
        if self.format == FileFormat.CSV:
            return ['Sheet1']
        elif self.format == FileFormat.EXCEL:
            return self.workbook.sheetnames
        else:
            return []
    
    def get_file_info(self) -> Dict[str, Any]:
        """
        获取文件信息
        
        Returns:
            Dict[str, Any]: 文件信息
        """
        info = {
            'file_path': self.file_path,
            'format': self.format.value,
            'file_size': os.path.getsize(self.file_path),
            'sheet_names': self.get_sheet_names()
        }
        
        if self.format == FileFormat.CSV:
            try:
                df = self.read_data()
                info.update({
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': df.columns.tolist()
                })
            except:
                pass
        elif self.format == FileFormat.EXCEL:
            try:
                # 获取第一个工作表的信息
                if self.workbook.sheetnames:
                    sheet_name = self.workbook.sheetnames[0]
                    worksheet = self.workbook[sheet_name]
                    info.update({
                        'max_row': worksheet.max_row,
                        'max_column': worksheet.max_column,
                        'used_range': worksheet.calculate_dimension()
                    })
            except:
                pass
        
        return info
    
    def close(self):
        """关闭文件"""
        if self.workbook:
            self.workbook.close()
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()


def detect_file_format(file_path: str) -> FileFormat:
    """
    检测文件格式
    
    Args:
        file_path (str): 文件路径
        
    Returns:
        FileFormat: 文件格式
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    file_extension = os.path.splitext(file_path)[1].lower()
    
    if file_extension in ['.xlsx', '.xls']:
        return FileFormat.EXCEL
    elif file_extension == '.csv':
        return FileFormat.CSV
    else:
        return FileFormat.UNKNOWN


def read_file(file_path: str, **kwargs) -> pd.DataFrame:
    """
    读取文件数据
    
    Args:
        file_path (str): 文件路径
        **kwargs: 其他参数
        
    Returns:
        pd.DataFrame: 读取的数据
    """
    with FileHandler(file_path) as handler:
        return handler.read_data(**kwargs)


def write_file(file_path: str, df: pd.DataFrame, **kwargs) -> None:
    """
    写入文件数据
    
    Args:
        file_path (str): 文件路径
        df (pd.DataFrame): 要写入的数据
        **kwargs: 其他参数
    """
    with FileHandler(file_path) as handler:
        handler.write_data(df, **kwargs)

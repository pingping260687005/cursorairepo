"""
Excel读取工具类

提供读取Excel文件的各种功能，包括：
- 读取整个工作表
- 读取指定范围的数据
- 读取多个工作表
- 支持不同的数据格式转换
"""

import pandas as pd
import openpyxl
from typing import Union, List, Dict, Optional, Any
import os


class ExcelReader:
    """Excel文件读取工具类"""
    
    def __init__(self, file_path: str):
        """
        初始化ExcelReader
        
        Args:
            file_path (str): Excel或CSV文件路径
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        self.file_path = file_path
        self.workbook = None
        self.is_csv = file_path.endswith('.csv')
        
        if not self.is_csv:
            self._load_workbook()
    
    def _load_workbook(self):
        """加载工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise Exception(f"无法加载Excel文件: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """
        获取所有工作表名称
        
        Returns:
            List[str]: 工作表名称列表
        """
        if self.is_csv:
            # CSV文件只有一个"工作表"
            return ['Sheet1']
        else:
            return self.workbook.sheetnames
    
    def read_sheet(self, sheet_name: Optional[str] = None, 
                   header_row: int = 0, 
                   use_cols: Optional[Union[str, List[int]]] = None,
                   encoding: str = 'utf-8',
                   delimiter: str = ',') -> pd.DataFrame:
        """
        读取指定工作表的数据
        
        Args:
            sheet_name (str, optional): 工作表名称，默认为第一个工作表
            header_row (int): 表头行号，默认为0
            use_cols (str or List[int], optional): 要读取的列，默认为所有列
            encoding (str): CSV文件编码，默认为utf-8
            delimiter (str): CSV分隔符，默认为逗号
            
        Returns:
            pd.DataFrame: 读取的数据
        """
        try:
            if self.file_path.endswith('.csv'):
                # CSV文件处理
                df = pd.read_csv(
                    self.file_path, 
                    header=header_row, 
                    usecols=use_cols,
                    encoding=encoding,
                    sep=delimiter
                )
            else:
                # Excel文件处理
                if sheet_name is None:
                    sheet_name = self.get_sheet_names()[0]
                
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    header=header_row,
                    usecols=use_cols
                )
            return df
        except Exception as e:
            raise Exception(f"读取工作表失败: {e}")
    
    def read_sheet_by_range(self, sheet_name: str, 
                           start_cell: str, 
                           end_cell: str) -> pd.DataFrame:
        """
        读取指定范围的数据
        
        Args:
            sheet_name (str): 工作表名称
            start_cell (str): 起始单元格，如'A1'
            end_cell (str): 结束单元格，如'C10'
            
        Returns:
            pd.DataFrame: 读取的数据
        """
        try:
            if self.is_csv:
                # CSV文件不支持范围读取，返回整个文件
                return self.read_sheet()
            else:
                # 使用openpyxl读取指定范围
                worksheet = self.workbook[sheet_name]
                data = []
                
                for row in worksheet[start_cell:end_cell]:
                    row_data = [cell.value for cell in row]
                    data.append(row_data)
                
                # 转换为DataFrame
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    return df
                else:
                    return pd.DataFrame()
                
        except Exception as e:
            raise Exception(f"读取指定范围数据失败: {e}")
    
    def read_all_sheets(self, header_row: int = 0) -> Dict[str, pd.DataFrame]:
        """
        读取所有工作表的数据
        
        Args:
            header_row (int): 表头行号，默认为0
            
        Returns:
            Dict[str, pd.DataFrame]: 工作表名称和数据的字典
        """
        sheets_data = {}
        
        for sheet_name in self.get_sheet_names():
            try:
                df = self.read_sheet(sheet_name, header_row)
                sheets_data[sheet_name] = df
            except Exception as e:
                print(f"读取工作表 '{sheet_name}' 失败: {e}")
                continue
        
        return sheets_data
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """
        获取工作表信息
        
        Args:
            sheet_name (str): 工作表名称
            
        Returns:
            Dict[str, Any]: 工作表信息
        """
        try:
            if self.is_csv:
                # CSV文件信息
                df = self.read_sheet()
                info = {
                    'sheet_name': sheet_name,
                    'max_row': len(df),
                    'max_column': len(df.columns),
                    'used_range': f"A1:{chr(ord('A') + len(df.columns) - 1)}{len(df)}",
                    'total_cells': len(df) * len(df.columns),
                    'used_cells': len(df) * len(df.columns)
                }
                return info
            else:
                worksheet = self.workbook[sheet_name]
                
                # 获取最大行数和列数
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # 获取有数据的范围
                used_range = worksheet.calculate_dimension()
                
                info = {
                    'sheet_name': sheet_name,
                    'max_row': max_row,
                    'max_column': max_col,
                    'used_range': used_range,
                    'total_cells': max_row * max_col,
                    'used_cells': len([cell for row in worksheet.iter_rows() for cell in row if cell.value is not None])
                }
                
                return info
            
        except Exception as e:
            raise Exception(f"获取工作表信息失败: {e}")
    
    def read_cell_value(self, sheet_name: str, cell_address: str) -> Any:
        """
        读取指定单元格的值
        
        Args:
            sheet_name (str): 工作表名称
            cell_address (str): 单元格地址，如'A1'
            
        Returns:
            Any: 单元格的值
        """
        try:
            if self.is_csv:
                # CSV文件不支持单元格地址读取，返回None
                return None
            else:
                worksheet = self.workbook[sheet_name]
                return worksheet[cell_address].value
        except Exception as e:
            raise Exception(f"读取单元格值失败: {e}")
    
    def read_column(self, sheet_name: str, column: str, 
                   start_row: int = 1, end_row: Optional[int] = None) -> List[Any]:
        """
        读取指定列的数据
        
        Args:
            sheet_name (str): 工作表名称
            column (str): 列名，如'A'或'B'
            start_row (int): 起始行号，默认为1
            end_row (int, optional): 结束行号，默认为最大行
            
        Returns:
            List[Any]: 列数据列表
        """
        try:
            if self.is_csv:
                # CSV文件不支持列地址读取，返回空列表
                return []
            else:
                worksheet = self.workbook[sheet_name]
                
                if end_row is None:
                    end_row = worksheet.max_row
                
                column_data = []
                for row in range(start_row, end_row + 1):
                    cell_value = worksheet[f"{column}{row}"].value
                    column_data.append(cell_value)
                
                return column_data
            
        except Exception as e:
            raise Exception(f"读取列数据失败: {e}")
    
    def read_row(self, sheet_name: str, row: int, 
                start_col: int = 1, end_col: Optional[int] = None) -> List[Any]:
        """
        读取指定行的数据
        
        Args:
            sheet_name (str): 工作表名称
            row (int): 行号
            start_col (int): 起始列号，默认为1
            end_col (int, optional): 结束列号，默认为最大列
            
        Returns:
            List[Any]: 行数据列表
        """
        try:
            if self.is_csv:
                # CSV文件不支持行地址读取，返回空列表
                return []
            else:
                worksheet = self.workbook[sheet_name]
                
                if end_col is None:
                    end_col = worksheet.max_column
                
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(cell_value)
                
                return row_data
            
        except Exception as e:
            raise Exception(f"读取行数据失败: {e}")
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
    
    def __enter__(self):
        """上下文管理器入口"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        self.close()

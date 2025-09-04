import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
import os
from pathlib import Path

class CSVCompareGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("CSVæ•°æ®æ¯”è¾ƒå·¥å…·")
        self.root.geometry("1024x768")
        
        # è®¾ç½®çª—å£æœ€å°å°ºå¯¸
        self.root.minsize(800, 600)
        
        # é…ç½®ä¸»é¢˜å’Œæ ·å¼
        self.style = ttk.Style()
        self.style.theme_use('clam')  # ä½¿ç”¨æ›´ç°ä»£çš„ä¸»é¢˜
        
        # é…ç½®å…¨å±€å­—ä½“
        default_font = ("å¾®è½¯é›…é»‘", 11)  # ä½¿ç”¨å¾®è½¯é›…é»‘ä½œä¸ºé»˜è®¤å­—ä½“ï¼Œè°ƒå¤§åˆ°11å·
        title_font = ("å¾®è½¯é›…é»‘", 24, "bold")  # æ ‡é¢˜å­—ä½“è°ƒå¤§åˆ°24å·
        button_font = ("å¾®è½¯é›…é»‘", 11)  # æŒ‰é’®å­—ä½“
        heading_font = ("å¾®è½¯é›…é»‘", 12, "bold")  # è¡¨æ ¼æ ‡é¢˜å­—ä½“
        self.root.option_add("*Font", default_font)
        
        # é…ç½®å…¨å±€æ ·å¼
        self.style.configure(".", font=default_font)
        self.style.configure("Title.TLabel", font=title_font, foreground="#2c3e50")
        self.style.configure("Nav.TButton", font=button_font, padding=10)
        self.style.configure("Action.TButton", 
                           font=button_font,
                           padding=(20, 10),
                           background="#3498db",
                           foreground="white")
        self.style.map("Action.TButton",
                      background=[("active", "#2980b9")])
        
        # è®¾ç½®Treeviewæ ·å¼
        self.style.configure("Treeview", 
                           font=default_font,
                           rowheight=35,  # å¢åŠ è¡Œé«˜ä»¥é€‚åº”æ›´å¤§çš„å­—ä½“
                           fieldbackground="#f8f9fa",
                           background="#ffffff")
        self.style.configure("Treeview.Heading", 
                           font=heading_font,  # ä½¿ç”¨æ›´å¤§çš„æ ‡é¢˜å­—ä½“
                           background="#f8f9fa")
        
        # è®¾ç½®Notebookæ ·å¼
        self.style.configure("TNotebook", padding=5)
        self.style.configure("TNotebook.Tab", 
                           padding=(20, 8),  # å¢åŠ æ ‡ç­¾é¡µå†…è¾¹è·
                           font=("å¾®è½¯é›…é»‘", 11, "bold"))  # æ ‡ç­¾é¡µå­—ä½“åŠ ç²—
        
        # åˆ›å»ºä¸»æ¡†æ¶ï¼Œä½¿ç”¨æ›´å¤§çš„å†…è¾¹è·
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # é…ç½®ä¸»çª—å£gridæƒé‡
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        
        # æ ‡é¢˜ï¼ˆå»é‡ï¼Œä»…ä¿ç•™ä¸€å¤„ï¼‰
        title = ttk.Label(self.main_frame, text="CSVæ•°æ®æ¯”è¾ƒå·¥å…·", style="Title.TLabel")
        title.grid(row=0, column=0, columnspan=2, pady=(0, 24))
        
        # åˆ›å»ºnotebookç”¨äºæ ‡ç­¾é¡µ
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # åˆ›å»ºä¸‰ä¸ªæ ‡ç­¾é¡µï¼Œä½¿ç”¨æ›´å¤§çš„å†…è¾¹è·
        self.mapping_frame = ttk.Frame(self.notebook, padding="20")
        self.upload_frame = ttk.Frame(self.notebook, padding="20")
        self.result_frame = ttk.Frame(self.notebook, padding="20")
        
        self.notebook.add(self.mapping_frame, text="1. å­—æ®µæ˜ å°„")
        self.notebook.add(self.upload_frame, text="2. æ–‡ä»¶ä¸Šä¼ ")
        self.notebook.add(self.result_frame, text="3. æ¯”è¾ƒç»“æœ")
        
        # åˆå§‹åŒ–å„ä¸ªé¡µé¢
        self.setup_mapping_page()
        self.setup_upload_page()
        self.setup_result_page()
        
        # æ•°æ®å­˜å‚¨
        self.mapping_data = {"field_mapping": {}, "key_fields": []}
        self.source_file = ""
        self.target_file = ""
        
        # åŠ è½½å·²æœ‰çš„mappingé…ç½®
        self.load_mapping()
    
    def setup_mapping_page(self):
        # å­—æ®µæ˜ å°„è¡¨æ ¼
        self.mapping_tree = ttk.Treeview(self.mapping_frame, columns=("source", "target", "desc"), show="headings")
        self.mapping_tree.heading("source", text="æ•°æ®1")
        self.mapping_tree.heading("target", text="æ•°æ®2")
        self.mapping_tree.heading("desc", text="æè¿°")
        
        # è®¾ç½®åˆ—å®½
        self.mapping_tree.column("source", width=150)
        self.mapping_tree.column("target", width=150)
        self.mapping_tree.column("desc", width=200)
        
        # æ·»åŠ æ»šåŠ¨æ¡
        scrollbar = ttk.Scrollbar(self.mapping_frame, orient="vertical", command=self.mapping_tree.yview)
        self.mapping_tree.configure(yscrollcommand=scrollbar.set)
        
        # ç»‘å®šåŒå‡»äº‹ä»¶
        self.mapping_tree.bind('<Double-1>', self.on_double_click)
        
        # æŒ‰é’®åŒºåŸŸ
        btn_frame = ttk.Frame(self.mapping_frame)
        ttk.Button(btn_frame, 
                  text="â• æ·»åŠ æ˜ å°„",
                  style="Action.TButton",
                  command=self.add_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame,
                  text="ğŸ—‘ åˆ é™¤æ˜ å°„",
                  style="Action.TButton",
                  command=self.delete_mapping).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame,
                  text="ğŸ’¾ ä¿å­˜é…ç½®",
                  style="Action.TButton",
                  command=self.save_mapping).pack(side=tk.LEFT, padx=5)
        
        # å…³é”®å­—æ®µåŒºåŸŸ
        key_frame = ttk.LabelFrame(self.mapping_frame, text="å…³é”®å­—æ®µ", padding="5")
        self.key_field_var = tk.StringVar()
        key_entry = ttk.Entry(key_frame, textvariable=self.key_field_var)
        ttk.Button(key_frame, text="æ·»åŠ å…³é”®å­—æ®µ", command=self.add_key_field).pack(side=tk.RIGHT, padx=5)
        ttk.Button(key_frame, text="åˆ é™¤é€‰ä¸­", command=self.delete_selected_key_field).pack(side=tk.RIGHT, padx=5)
        key_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # å…³é”®å­—æ®µåˆ—è¡¨
        self.key_listbox = tk.Listbox(key_frame, height=4)
        key_scrollbar = ttk.Scrollbar(key_frame, orient="vertical", command=self.key_listbox.yview)
        self.key_listbox.configure(yscrollcommand=key_scrollbar.set)
        # åŒå‡»åˆ é™¤å…³é”®å­—æ®µ
        self.key_listbox.bind('<Double-1>', lambda e: self.delete_selected_key_field())
        
        # å¸ƒå±€
        self.mapping_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        btn_frame.grid(row=1, column=0, columnspan=2, pady=10)
        key_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        self.key_listbox.pack(fill=tk.X, expand=True, pady=5)
        key_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # é…ç½®æƒé‡
        self.mapping_frame.columnconfigure(0, weight=1)
        self.mapping_frame.rowconfigure(0, weight=1)
    
    def setup_upload_page(self):
        # æ–‡ä»¶1é€‰æ‹©
        file1_frame = ttk.LabelFrame(self.upload_frame, text="æ•°æ®æ–‡ä»¶1", padding="5")
        self.file1_var = tk.StringVar()
        ttk.Entry(file1_frame, textvariable=self.file1_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(file1_frame, text="é€‰æ‹©æ–‡ä»¶", command=lambda: self.select_file("source")).pack(side=tk.RIGHT)
        
        # æ–‡ä»¶2é€‰æ‹©
        file2_frame = ttk.LabelFrame(self.upload_frame, text="æ•°æ®æ–‡ä»¶2", padding="5")
        self.file2_var = tk.StringVar()
        ttk.Entry(file2_frame, textvariable=self.file2_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(file2_frame, text="é€‰æ‹©æ–‡ä»¶", command=lambda: self.select_file("target")).pack(side=tk.RIGHT)
        
        # æ–‡ä»¶ä¸Šä¼ è¯´æ˜
        instruction = ttk.Label(self.upload_frame,
                              text="è¯·é€‰æ‹©è¦æ¯”è¾ƒçš„CSVæ–‡ä»¶",
                              font=("Microsoft YaHei UI", 12, "bold"),
                              foreground="#2c3e50")
        
        # æ¯”è¾ƒæŒ‰é’®
        compare_btn = ttk.Button(self.upload_frame,
                               text="ğŸ” å¼€å§‹æ¯”è¾ƒ",
                               style="Action.TButton",
                               command=self.compare_files)
        
        # å¸ƒå±€
        instruction.pack(pady=(0, 20))
        file1_frame.pack(fill=tk.X, pady=10)
        file2_frame.pack(fill=tk.X, pady=10)
        compare_btn.pack(pady=30)
    
    def setup_result_page(self):
        # åˆ›å»ºç»“æœæ ‡ç­¾å’Œç»Ÿè®¡ä¿¡æ¯åŒºåŸŸ
        info_frame = ttk.Frame(self.result_frame)
        info_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.diff_count_label = ttk.Label(info_frame, text="å·®å¼‚ç»Ÿè®¡ï¼š")
        self.diff_count_label.pack(side=tk.LEFT, padx=5)
        
        # å›¾ä¾‹è¯´æ˜
        legend_frame = ttk.Frame(info_frame, padding=(5, 5))
        legend_frame.pack(side=tk.RIGHT, padx=5)
        
        # ä½¿ç”¨æ›´ç°ä»£çš„é¢œè‰²æ–¹æ¡ˆ
        legend_style = {"font": ("å¾®è½¯é›…é»‘", 10),
                       "style": "Legend.TLabel"}
                       
        # é…ç½®å›¾ä¾‹æ ‡ç­¾æ ·å¼
        self.style.configure("Legend.TLabel",
                           font=("å¾®è½¯é›…é»‘", 10),
                           background=self.style.lookup("TFrame", "background"))
        
        # åˆ›å»ºé—´éš”å°ä¸€äº›çš„å›¾ä¾‹æ ‡ç­¾
        ttk.Label(legend_frame, text="å›¾ä¾‹è¯´æ˜ï¼š", **legend_style).pack(side=tk.LEFT)
        ttk.Label(legend_frame, text="â¬¤ ä»…åœ¨æ•°æ®1ä¸­", foreground="#e74c3c", **legend_style).pack(side=tk.LEFT, padx=(10, 5))
        ttk.Label(legend_frame, text="â¬¤ ä»…åœ¨æ•°æ®2ä¸­", foreground="#27ae60", **legend_style).pack(side=tk.LEFT, padx=5)
        ttk.Label(legend_frame, text="â¬¤ æ•°æ®ä¸ä¸€è‡´", foreground="#f1c40f", **legend_style).pack(side=tk.LEFT, padx=5)
        
        # ç»“æœå±•ç¤ºè¡¨æ ¼
        self.result_tree = ttk.Treeview(self.result_frame, show="headings")
        
        # è®¾ç½®è¡¨æ ¼æ ·å¼
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))
        # ç»“æœè¡¨æ ¼æ–‘é©¬çº¹æ ‡ç­¾
        self.result_tree.tag_configure('oddrow', background="#fafafa")
        self.result_tree.tag_configure('evenrow', background="#f3f4f6")
        
        # æ·»åŠ æ»šåŠ¨æ¡
        y_scrollbar = ttk.Scrollbar(self.result_frame, orient="vertical", command=self.result_tree.yview)
        x_scrollbar = ttk.Scrollbar(self.result_frame, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # å¯¼å‡ºæŒ‰é’®å’Œæ“ä½œåŒºåŸŸ
        btn_frame = ttk.Frame(self.result_frame)
        ttk.Button(btn_frame,
                  text="ğŸ“Š å¯¼å‡ºExcel",
                  style="Action.TButton",
                  command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame,
                  text="ğŸ“‹ å¤åˆ¶åˆ°å‰ªè´´æ¿",
                  style="Action.TButton",
                  command=self.copy_to_clipboard).pack(side=tk.LEFT, padx=5)
        
        # å¸ƒå±€
        self.result_tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        y_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        x_scrollbar.grid(row=2, column=0, sticky=(tk.W, tk.E))
        btn_frame.grid(row=3, column=0, columnspan=2, pady=10)
        
        # é…ç½®æƒé‡
        self.result_frame.columnconfigure(0, weight=1)
        self.result_frame.rowconfigure(1, weight=1)
        
        # ç»‘å®šåŒå‡»äº‹ä»¶ä»¥æŸ¥çœ‹è¯¦ç»†å·®å¼‚
        self.result_tree.bind("<Double-1>", self.show_diff_detail)
    
    def create_edit_dialog(self, title="æ·»åŠ å­—æ®µæ˜ å°„", values=None):
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("400x300")  # å¢åŠ çª—å£å°ºå¯¸
        dialog.transient(self.root)  # è®¾ç½®ä¸ºä¸»çª—å£çš„ä¸´æ—¶çª—å£
        dialog.grab_set()  # æ¨¡æ€çª—å£
        # dialog.resizable(False, False)  # ç¦æ­¢è°ƒæ•´å¤§å°
        
        # è¾“å…¥æ¡†
        source_var = tk.StringVar(value=values[0] if values else "")
        target_var = tk.StringVar(value=values[1] if values else "")
        desc_var = tk.StringVar(value=values[2] if values else "")
        
        # åˆ›å»ºè¡¨å•æ¡†æ¶
        form_frame = ttk.Frame(dialog, padding="10")
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # åˆ›å»ºæ ·å¼
        style = ttk.Style()
        style.configure("Field.TLabel", font=("å¾®è½¯é›…é»‘", 11))
        style.configure("Field.TEntry", padding=5)
        
        # æ•°æ®1å­—æ®µ
        ttk.Label(form_frame, text="æ•°æ®1å­—æ®µ:", style="Field.TLabel").pack(anchor=tk.W, pady=(0, 5))
        source_entry = ttk.Entry(form_frame, textvariable=source_var, style="Field.TEntry", width=40)
        source_entry.pack(fill=tk.X, pady=(0, 15))
        
        # æ•°æ®2å­—æ®µ
        ttk.Label(form_frame, text="æ•°æ®2å­—æ®µ:", style="Field.TLabel").pack(anchor=tk.W, pady=(0, 5))
        target_entry = ttk.Entry(form_frame, textvariable=target_var, style="Field.TEntry", width=40)
        target_entry.pack(fill=tk.X, pady=(0, 15))
        
        # æè¿°å­—æ®µ
        ttk.Label(form_frame, text="æè¿°:", style="Field.TLabel").pack(anchor=tk.W, pady=(0, 5))
        desc_entry = ttk.Entry(form_frame, textvariable=desc_var, style="Field.TEntry", width=40)
        desc_entry.pack(fill=tk.X, pady=(0, 15))
        
        result = {"submit": False, "values": None}
        
        def submit():
            source = source_var.get().strip()
            target = target_var.get().strip()
            desc = desc_var.get().strip()
            if source and target:
                result["submit"] = True
                result["values"] = (source, target, desc)
                dialog.destroy()
            else:
                messagebox.showwarning("è­¦å‘Š", "æ•°æ®1å­—æ®µå’Œæ•°æ®2å­—æ®µä¸èƒ½ä¸ºç©ºï¼", parent=dialog)
        
        # æŒ‰é’®åŒºåŸŸ
        # æŒ‰é’®æ ·å¼
        style.configure("Action.TButton", padding=10)
        
        btn_frame = ttk.Frame(form_frame)
        btn_frame.pack(fill=tk.X, pady=(20, 0))
        ttk.Button(btn_frame, text="ç¡®å®š", command=submit, style="Action.TButton", width=15).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="å–æ¶ˆ", command=dialog.destroy, style="Action.TButton", width=15).pack(side=tk.RIGHT, padx=5)
        
        # è®¾ç½®ç„¦ç‚¹å¹¶ç­‰å¾…çª—å£å…³é—­
        source_entry.focus_set()
        dialog.wait_window()
        return result
    
    def add_mapping(self):
        result = self.create_edit_dialog()
        if result["submit"] and result["values"]:
            self.mapping_tree.insert("", tk.END, values=result["values"])
            self.save_mapping()  # è‡ªåŠ¨ä¿å­˜æ›´æ”¹
    
    def edit_mapping(self, item):
        current_values = self.mapping_tree.item(item)["values"]
        result = self.create_edit_dialog("ç¼–è¾‘å­—æ®µæ˜ å°„", current_values)
        if result["submit"] and result["values"]:
            self.mapping_tree.item(item, values=result["values"])
            self.save_mapping()  # è‡ªåŠ¨ä¿å­˜æ›´æ”¹
    
    def on_double_click(self, event):
        item = self.mapping_tree.selection()[0]
        self.edit_mapping(item)
    
    def delete_mapping(self):
        selected = self.mapping_tree.selection()
        if selected:
            if messagebox.askyesno("ç¡®è®¤", "ç¡®å®šè¦åˆ é™¤é€‰ä¸­çš„æ˜ å°„å—ï¼Ÿ"):
                self.mapping_tree.delete(selected)
                self.save_mapping()  # è‡ªåŠ¨ä¿å­˜æ›´æ”¹
    
    def add_key_field(self):
        field = self.key_field_var.get()
        if field:
            self.key_listbox.insert(tk.END, field)
            self.key_field_var.set("")
    
    def load_mapping(self):
        try:
            if os.path.exists("mapping.csv"):
                df = pd.read_csv("mapping.csv")
                for _, row in df.iterrows():
                    # ä½¿ç”¨source1å’Œsource2åˆ—ä½œä¸ºæ˜ å°„æºå’Œç›®æ ‡
                    source = row.get("source1", "")
                    target = row.get("source2", "")
                    desc = row.get("desc", "")
                    if source and target:  # åªæ·»åŠ éç©ºçš„æ˜ å°„
                        self.mapping_tree.insert("", tk.END, values=(source, target, desc))
                
                # åŠ è½½å…³é”®å­—æ®µ (is_keyä¸ºyesçš„å­—æ®µ)
                key_fields = df[df["is_key"] == "yes"]["source1"].tolist()
                for field in key_fields:
                    if field:  # åªæ·»åŠ éç©ºçš„å…³é”®å­—æ®µ
                        self.key_listbox.insert(tk.END, field)
        except Exception as e:
            messagebox.showerror("é”™è¯¯", f"åŠ è½½mapping.csvå¤±è´¥: {str(e)}")
            print(f"é”™è¯¯è¯¦æƒ…: {str(e)}")  # åœ¨æ§åˆ¶å°æ‰“å°è¯¦ç»†é”™è¯¯ä¿¡æ¯
    
    def save_mapping(self):
        try:
            # æ”¶é›†æ˜ å°„æ•°æ®
            mapping_data = []
            for item in self.mapping_tree.get_children():
                values = self.mapping_tree.item(item)["values"]
                key_fields = list(self.key_listbox.get(0, tk.END))
                mapping_data.append({
                    "source1": values[0],
                    "source2": values[1],
                    "desc": values[2] if len(values) > 2 else "",
                    "is_key": "yes" if values[0] in key_fields else "no"
                })
            
            # ä¿å­˜åˆ°CSV
            mapping_df = pd.DataFrame(mapping_data)
            mapping_df.to_csv("mapping.csv", index=False)
            messagebox.showinfo("æˆåŠŸ", "é…ç½®å·²ä¿å­˜åˆ°mapping.csv")
        except Exception as e:
            messagebox.showerror("é”™è¯¯", f"ä¿å­˜é…ç½®å¤±è´¥: {str(e)}")
    
    def select_file(self, file_type):
        filename = filedialog.askopenfilename(
            title=f"é€‰æ‹©{'æ•°æ®æ–‡ä»¶1' if file_type == 'source' else 'æ•°æ®æ–‡ä»¶2'}",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            if file_type == "source":
                self.source_file = filename
                self.file1_var.set(filename)
            else:
                self.target_file = filename
                self.file2_var.set(filename)
    
    def compare_files(self):
        if not (self.source_file and self.target_file):
            messagebox.showwarning("è­¦å‘Š", "è¯·å…ˆé€‰æ‹©è¦æ¯”è¾ƒçš„æ–‡ä»¶")
            return
        
        try:
            # è¯»å–CSVæ–‡ä»¶
            df1 = pd.read_csv(self.source_file)
            df2 = pd.read_csv(self.target_file)
            
            # è·å–æ˜ å°„å…³ç³»
            field_mapping = {}
            non_key_fields = []  # è®°å½•éå…³é”®å­—æ®µ
            for item in self.mapping_tree.get_children():
                values = self.mapping_tree.item(item)["values"]
                source_field, target_field = values[0], values[1]
                field_mapping[target_field] = source_field
                non_key_fields.append(source_field)
            
            # è·å–å…³é”®å­—æ®µ
            key_fields = list(self.key_listbox.get(0, tk.END))
            if not key_fields:
                messagebox.showerror("é”™è¯¯", "è¯·è‡³å°‘è®¾ç½®ä¸€ä¸ªå…³é”®å­—æ®µ")
                return
                
            # ä»éå…³é”®å­—æ®µä¸­ç§»é™¤å…³é”®å­—æ®µ
            non_key_fields = [f for f in non_key_fields if f not in key_fields]
            
            # é‡å‘½ådf2çš„åˆ—ä»¥åŒ¹é…df1
            df2_renamed = df2.rename(columns=field_mapping)
            
            # å‡†å¤‡ç»“æœåˆ—
            result_columns = key_fields.copy()
            for field in non_key_fields:
                result_columns.extend([f"{field}_æ•°æ®1", f"{field}_æ•°æ®2"])
            
            # åˆå¹¶æ•°æ®
            merged = pd.merge(df1, df2_renamed, on=key_fields, how='outer', 
                            suffixes=('_æ•°æ®1', '_æ•°æ®2'), indicator=True)
            
            # æ¸…é™¤ç°æœ‰ç»“æœ
            for col in self.result_tree["columns"]:
                self.result_tree.heading(col, text="")
            
            # è®¾ç½®æ–°çš„åˆ—
            self.result_tree["columns"] = tuple(result_columns)
            
            # è®¾ç½®åˆ—æ ‡é¢˜å’Œå®½åº¦
            for col in result_columns:
                display_name = col.replace("_æ•°æ®1", "(æ•°æ®1)").replace("_æ•°æ®2", "(æ•°æ®2)")
                self.result_tree.heading(col, text=display_name)
                self.result_tree.column(col, width=120, anchor="center")
            
            # æ˜¾ç¤ºæ•°æ®å¹¶è®¾ç½®è¡Œé¢œè‰²
            only_in_source = 0
            only_in_target = 0
            diff_values = 0
            
            # å¤„ç†æ¯ä¸€è¡Œæ•°æ®
            diff_rows = []  # å­˜å‚¨æœ‰å·®å¼‚çš„è¡Œæ•°æ®
            diff_tags = []  # å­˜å‚¨å¯¹åº”çš„å·®å¼‚æ ‡ç­¾
            
            for index, row in merged.iterrows():
                merge_status = row["_merge"]
                values = []
                
                # æ·»åŠ å…³é”®å­—æ®µçš„å€¼
                for key in key_fields:
                    values.append(str(row[key]))
                
                # æ·»åŠ éå…³é”®å­—æ®µçš„æ¯”è¾ƒå€¼
                has_diff = False
                for field in non_key_fields:
                    val1 = row.get(f"{field}_æ•°æ®1", "")
                    val2 = row.get(f"{field}_æ•°æ®2", "")
                    values.extend([str(val1), str(val2)])
                    
                    # æ£€æŸ¥æ˜¯å¦æœ‰å·®å¼‚
                    if pd.notna(val1) and pd.notna(val2) and str(val1) != str(val2):
                        has_diff = True
                
                # åªæ·»åŠ æœ‰å·®å¼‚çš„è¡Œ
                if merge_status == "left_only":  # ä»…åœ¨æ•°æ®1ä¸­
                    diff_rows.append(values)
                    diff_tags.append("left_only")
                    only_in_source += 1
                elif merge_status == "right_only":  # ä»…åœ¨æ•°æ®2ä¸­
                    diff_rows.append(values)
                    diff_tags.append("right_only")
                    only_in_target += 1
                elif has_diff:  # å€¼ä¸ä¸€è‡´
                    diff_rows.append(values)
                    diff_tags.append("diff_values")
                    diff_values += 1
            
            # é…ç½®æ ‡ç­¾æ ·å¼ - ä½¿ç”¨æ›´æŸ”å’Œçš„é¢œè‰²
            self.result_tree.tag_configure("left_only", background="#fbe9e7")  # æ›´æŸ”å’Œçš„çº¢è‰²
            self.result_tree.tag_configure("right_only", background="#e8f5e9")  # æ›´æŸ”å’Œçš„ç»¿è‰²
            self.result_tree.tag_configure("diff_values", background="#fff8e1")  # æ›´æŸ”å’Œçš„é»„è‰²
            
            # æ˜¾ç¤ºå·®å¼‚è¡Œ
            for idx, (values, tag) in enumerate(zip(diff_rows, diff_tags)):
                zebra_tag = 'oddrow' if idx % 2 == 0 else 'evenrow'
                item_id = self.result_tree.insert("", tk.END, values=values)
                self.result_tree.item(item_id, tags=(tag, zebra_tag))
                
            # ä¿å­˜å·®å¼‚æ•°æ®ä¾›å¯¼å‡ºä½¿ç”¨
            self.diff_data = list(zip(diff_rows, diff_tags))
            
            # æ›´æ–°ç»Ÿè®¡ä¿¡æ¯
            total_diff = only_in_source + only_in_target + diff_values
            stats_text = f"å·®å¼‚ç»Ÿè®¡ï¼šæ€»å·®å¼‚æ•° {total_diff} | ä»…åœ¨æ•°æ®1ä¸­ {only_in_source} | ä»…åœ¨æ•°æ®2ä¸­ {only_in_target} | å€¼ä¸ä¸€è‡´ {diff_values}"
            self.diff_count_label.config(text=stats_text)
            
            # åˆ‡æ¢åˆ°ç»“æœé¡µé¢
            self.notebook.select(2)
            
        except Exception as e:
            messagebox.showerror("é”™è¯¯", f"æ¯”è¾ƒæ–‡ä»¶å¤±è´¥: {str(e)}")
            if hasattr(self, 'diff_count_label'):
                self.diff_count_label.config(text="å·®å¼‚ç»Ÿè®¡ï¼šå‘ç”Ÿé”™è¯¯ï¼Œè¯·æ£€æŸ¥è¾“å…¥æ–‡ä»¶ä¸æ˜ å°„é…ç½®")

    def delete_selected_key_field(self):
        """åˆ é™¤é€‰ä¸­çš„å…³é”®å­—æ®µ"""
        selection = self.key_listbox.curselection()
        if not selection:
            return
        for index in reversed(selection):
            self.key_listbox.delete(index)
            
    def has_value_differences(self, row, key_fields):
        """æ£€æŸ¥éå…³é”®å­—æ®µæ˜¯å¦æœ‰å·®å¼‚"""
        try:
            for col in row.index:
                if col not in key_fields and col != "_merge":
                    x_col = col + "_x" if col + "_x" in row.index else col
                    y_col = col + "_y" if col + "_y" in row.index else col
                    if x_col != y_col and pd.notna(row.get(x_col)) and pd.notna(row.get(y_col)):
                        if str(row.get(x_col)) != str(row.get(y_col)):
                            return True
            return False
        except Exception:
            return False
            
    def copy_to_clipboard(self):
        """å°†é€‰ä¸­çš„è¡Œå¤åˆ¶åˆ°å‰ªè´´æ¿"""
        selected_items = self.result_tree.selection()
        if not selected_items:
            messagebox.showinfo("æç¤º", "è¯·å…ˆé€‰æ‹©è¦å¤åˆ¶çš„è¡Œ")
            return
            
        text = []
        # æ·»åŠ è¡¨å¤´
        headers = [self.result_tree.heading(col)["text"] for col in self.result_tree["columns"]]
        text.append("\t".join(headers))
        
        # æ·»åŠ é€‰ä¸­çš„è¡Œ
        for item in selected_items:
            values = self.result_tree.item(item)["values"]
            text.append("\t".join(map(str, values)))
            
        # å¤åˆ¶åˆ°å‰ªè´´æ¿
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(text))
        messagebox.showinfo("æˆåŠŸ", "å·²å¤åˆ¶åˆ°å‰ªè´´æ¿")
            
    def show_diff_detail(self, event):
        """æ˜¾ç¤ºè¡Œå·®å¼‚è¯¦æƒ…"""
        item = self.result_tree.identify_row(event.y)
        if not item:
            return
            
        values = self.result_tree.item(item)["values"]
        tags = self.result_tree.item(item)["tags"]
        
        if not tags:
            return
            
        diff_type = {
            "left_only": "æ­¤è¡Œä»…åœ¨æ•°æ®1ä¸­å­˜åœ¨",
            "right_only": "æ­¤è¡Œä»…åœ¨æ•°æ®2ä¸­å­˜åœ¨",
            "diff_values": "æ­¤è¡Œæ•°æ®ä¸ä¸€è‡´"
        }.get(tags[0], "")
        
        if diff_type:
            # åˆ›å»ºè¯¦æƒ…çª—å£
            detail = tk.Toplevel(self.root)
            detail.title("å·®å¼‚è¯¦æƒ…")
            detail.geometry("600x500")
            detail.transient(self.root)
            
            # åˆ›å»ºFrameç”¨äºç»„ç»‡å†…å®¹
            main_frame = ttk.Frame(detail, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # æ˜¾ç¤ºå·®å¼‚ç±»å‹
            ttk.Label(main_frame, text=diff_type, font=("Helvetica", 12, "bold")).pack(anchor=tk.W, pady=(0, 15))
            
            # åˆ›å»ºè¡¨æ ¼æ˜¾ç¤ºå·®å¼‚
            tree = ttk.Treeview(main_frame, columns=("å­—æ®µ", "æ•°æ®1", "æ•°æ®2", "çŠ¶æ€"), show="headings", height=10)
            tree.heading("å­—æ®µ", text="å­—æ®µ")
            tree.heading("æ•°æ®1", text="æ•°æ®1å€¼")
            tree.heading("æ•°æ®2", text="æ•°æ®2å€¼")
            tree.heading("çŠ¶æ€", text="çŠ¶æ€")
            
            # è®¾ç½®åˆ—å®½
            tree.column("å­—æ®µ", width=100)
            tree.column("æ•°æ®1", width=150)
            tree.column("æ•°æ®2", width=150)
            tree.column("çŠ¶æ€", width=100)
            
            # æ·»åŠ æ»šåŠ¨æ¡
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            
            # å¸ƒå±€
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # è·å–æ‰€æœ‰å­—æ®µ
            columns = self.result_tree["columns"]
            key_fields = list(self.key_listbox.get(0, tk.END))
            
            # æ˜¾ç¤ºæ•°æ®
            idx = 0
            # é¦–å…ˆæ˜¾ç¤ºå…³é”®å­—æ®µ
            for key in key_fields:
                val = values[idx]
                tree.insert("", tk.END, values=(key, val, val, "å…³é”®å­—æ®µ"))
                idx += 1
            
            # ç„¶åæ˜¾ç¤ºå…¶ä»–å­—æ®µ
            while idx < len(values):
                field_name = columns[idx].replace("_æ•°æ®1", "")
                if field_name + "_æ•°æ®2" in columns:
                    val1 = values[idx]
                    val2 = values[idx + 1]
                    status = "ä¸ä¸€è‡´" if val1 != val2 and pd.notna(val1) and pd.notna(val2) else "ä¸€è‡´"
                    tree.insert("", tk.END, values=(field_name, val1, val2, status),
                              tags=("diff",) if status == "ä¸ä¸€è‡´" else ())
                    idx += 2
                else:
                    idx += 1
            
            # è®¾ç½®ä¸ä¸€è‡´è¡Œçš„é¢œè‰²
            tree.tag_configure("diff", background="#fff9c4")
            
            # æ·»åŠ å…³é—­æŒ‰é’®
            ttk.Button(main_frame, text="å…³é—­", command=detail.destroy).pack(pady=(15, 0))
    
    def export_excel(self):
        if not hasattr(self, 'diff_data') or not self.diff_data:
            messagebox.showwarning("è­¦å‘Š", "æ²¡æœ‰å·®å¼‚æ•°æ®å¯å¯¼å‡º")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        # è·Ÿè¸ªåŒ…å«å·®å¼‚çš„å•å…ƒæ ¼
        diff_cells = []

        if filename:
            try:
                # åˆ›å»ºExcel writerå¯¹è±¡
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # è·å–å…³é”®å­—æ®µå’Œåˆ—ä¿¡æ¯
                    key_fields = list(self.key_listbox.get(0, tk.END))
                    columns = self.result_tree["columns"]
                    data = []
                    diff_cells = []  # å­˜å‚¨éœ€è¦æ ‡è®°çš„å•å…ƒæ ¼ä½ç½®
                    
                    # ä»ä¿å­˜çš„å·®å¼‚æ•°æ®ä¸­è·å–æ•°æ®
                    for values, tag in self.diff_data:
                        data.append(values)
                        
                        # æ£€æŸ¥éå…³é”®å­—æ®µçš„å·®å¼‚
                        if tag == "diff_values":
                            row_idx = len(data)  # å½“å‰è¡Œå·ï¼ˆè€ƒè™‘åˆ°åé¢Excelä¸­çš„æ ‡é¢˜è¡Œï¼Œå®é™…è¡Œå·ä¼š+1ï¼‰
                            idx = len(key_fields)  # ä»éå…³é”®å­—æ®µå¼€å§‹æ£€æŸ¥
                            
                            while idx < len(values):
                                if idx + 1 < len(values):  # ç¡®ä¿æœ‰æˆå¯¹çš„å€¼
                                    val1 = str(values[idx])
                                    val2 = str(values[idx + 1])
                                    if val1 != val2:  # å¦‚æœå€¼ä¸åŒï¼Œè®°å½•å•å…ƒæ ¼ä½ç½®
                                        diff_cells.append((row_idx, idx))
                                        diff_cells.append((row_idx, idx + 1))
                                idx += 2
                    
                    # åˆ›å»ºDataFrame
                    df = pd.DataFrame(data, columns=[col.replace("_æ•°æ®1", "(æ•°æ®1)").replace("_æ•°æ®2", "(æ•°æ®2)") 
                                                   for col in columns])
                    
                    # å†™å…¥æ•°æ®
                    df.to_excel(writer, sheet_name='æ¯”è¾ƒç»“æœ', index=False)
                    
                    # è·å–å·¥ä½œè¡¨å¯¹è±¡
                    worksheet = writer.sheets['æ¯”è¾ƒç»“æœ']
                    
                    # ä»openpyxl.styleså¯¼å…¥æ‰€éœ€çš„æ ·å¼
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

                    # åœ¨å†™å…¥æ•°æ®åæ£€æŸ¥å·®å¼‚å•å…ƒæ ¼
                    for row_idx, (values, tag) in enumerate(self.diff_data, start=2):  # ä»ç¬¬2è¡Œå¼€å§‹ï¼ˆè·³è¿‡æ ‡é¢˜ï¼‰
                        if tag == "diff_values":
                            idx = len(key_fields)  # ä»éå…³é”®å­—æ®µå¼€å§‹æ£€æŸ¥
                            while idx < len(values):
                                if idx + 1 < len(values):  # ç¡®ä¿æœ‰æˆå¯¹çš„å€¼
                                    val1 = str(values[idx])
                                    val2 = str(values[idx + 1])
                                    if val1 != val2:  # å¦‚æœå€¼ä¸åŒï¼Œè®°å½•å•å…ƒæ ¼ä½ç½®
                                        diff_cells.append((row_idx, idx + 1))
                                        diff_cells.append((row_idx, idx + 2))
                                idx += 2
                    
                    # å®šä¹‰æ ·å¼
                    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    
                    # è®¾ç½®åˆ—å®½
                    for column in worksheet.columns:
                        max_length = 0
                        column = list(column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    # è®¾ç½®æ ‡é¢˜è¡Œæ ·å¼
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # è®¾ç½®æ•°æ®è¡Œæ ·å¼
                    for row_idx, (values, tag) in enumerate(self.diff_data, start=2):  # ä»ç¬¬2è¡Œå¼€å§‹ï¼ˆè·³è¿‡æ ‡é¢˜ï¼‰
                        # è®¾ç½®è¡ŒåŸºæœ¬æ ·å¼
                        for cell in worksheet[row_idx]:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # è®¾ç½®æ•´è¡Œçš„åŸºæœ¬é¢œè‰²
                        if tag in ["left_only", "right_only"]:
                            fill = {
                                "left_only": red_fill,
                                "right_only": green_fill
                            }.get(tag)
                            
                            for cell in worksheet[row_idx]:
                                cell.fill = fill
                        
                        # æ ‡è®°å·®å¼‚å€¼çš„å•å…ƒæ ¼
                        for cell_idx, cell in enumerate(worksheet[row_idx]):
                            if (row_idx, cell_idx + 1) in diff_cells:  # +1 å› ä¸ºExcelåˆ—å·ä»1å¼€å§‹
                                cell.fill = yellow_fill
                    
                    # æ·»åŠ è¾¹æ¡†
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.border = thin_border
                    
                    # å†»ç»“é¦–è¡Œ
                    worksheet.freeze_panes = 'A2'
                    
                messagebox.showinfo("æˆåŠŸ", f"æ•°æ®å·²å¯¼å‡ºåˆ°: {filename}\n\nåŒ…å«ä»¥ä¸‹ç‰¹æ€§ï¼š\n- å·®å¼‚è¡Œé¢œè‰²æ ‡è®°\n- è‡ªåŠ¨è°ƒæ•´åˆ—å®½\n- å±…ä¸­å¯¹é½\n- å†»ç»“æ ‡é¢˜è¡Œ\n- è¾¹æ¡†ç¾åŒ–")
            except Exception as e:
                messagebox.showerror("é”™è¯¯", f"å¯¼å‡ºå¤±è´¥: {str(e)}")

def main():
    root = tk.Tk()
    app = CSVCompareGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()

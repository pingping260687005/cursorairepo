"""
CSV数据比较路由
"""
from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import io
import logging
from typing import Dict, List, Tuple, Any
import tempfile
import os
from datetime import datetime
import openpyxl.styles
import csv
import json

# 创建蓝图
data_compare_bp = Blueprint('data_compare', __name__, url_prefix='/data')

# 配置日志
logger = logging.getLogger(__name__)

@data_compare_bp.route('/compare', methods=['GET'])
def show_compare_form():
    """显示CSV比较表单页面"""
    try:
        with open('templates/compare_form.html', 'r', encoding='utf-8') as f:
            html_content = f.read()
        return html_content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except FileNotFoundError:
        return jsonify({
            'status': 'error',
            'message': 'HTML template not found',
            'endpoint': '/data/compare'
        }), 404

@data_compare_bp.route('/compare', methods=['POST'])
def compare_csv():
    """
    CSV数据比较端点
    
    请求参数:
    - source_csv: 源CSV文件
    - target_csv: 目标CSV文件  
    - field_mapping: 字段映射关系 (JSON格式)
    - key_fields: 关键字段列表 (用于关联记录)
    
    Returns:
        Excel文件: 包含比较结果的Excel文件
    """
    try:
        # 检查是否有文件上传
        if 'source_csv' not in request.files or 'target_csv' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'Both source_csv and target_csv files are required',
                'endpoint': '/data/compare'
            }), 400
        
        source_file = request.files['source_csv']
        target_file = request.files['target_csv']
        
        # 检查文件类型
        if not source_file.filename.endswith('.csv') or not target_file.filename.endswith('.csv'):
            return jsonify({
                'status': 'error',
                'message': 'Both files must be CSV format',
                'endpoint': '/data/compare'
            }), 400
        
        # 获取字段映射和关键字段
        mapping_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'mapping.csv')
        import pandas as pd
        if os.path.exists(mapping_file):
            df = pd.read_csv(mapping_file, dtype=str).fillna('')
            field_mapping = dict(zip(df['source1'], df['source2']))
            key_fields = df.loc[df['is_key'].str.lower() == 'yes', 'source1'].tolist()
        else:
            field_mapping = {}
            key_fields = []
        
        # 读取CSV文件
        source_df = pd.read_csv(source_file)
        target_df = pd.read_csv(target_file)
        
        logger.info(f"Source CSV loaded: {len(source_df)} rows, {len(source_df.columns)} columns")
        logger.info(f"Target CSV loaded: {len(target_df)} rows, {len(target_df.columns)} columns")
        
        # 执行数据比较
        comparison_result = compare_dataframes(source_df, target_df, field_mapping, key_fields)
        
        # 生成Excel报告
        excel_file = generate_excel_report(comparison_result)
        
        # 返回Excel文件
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'csv_comparison_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error in CSV comparison: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e),
            'endpoint': '/data/compare'
        }), 500

def compare_dataframes(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                      field_mapping: Dict[str, str], key_fields: List[str]) -> Dict[str, Any]:
    """
    比较两个DataFrame
    
    Args:
        source_df: 源数据框
        target_df: 目标数据框
        field_mapping: 字段映射关系
        key_fields: 关键字段列表
        
    Returns:
        比较结果字典
    """
    result = {
        'data_loss': [],
        'value_diff': [],
        'summary': {}
    }
    
    # 如果没有字段映射，使用交集字段
    if not field_mapping:
        common_fields = list(set(source_df.columns) & set(target_df.columns))
        field_mapping = {field: field for field in common_fields}
    
    # 如果没有关键字段，使用第一个字段作为关键字段
    if not key_fields:
        key_fields = [list(source_df.columns)[0]]
    
    logger.info(f"Field mapping: {field_mapping}")
    logger.info(f"Key fields: {key_fields}")
    
    # 应用字段映射到目标表
    if field_mapping:
        # 创建字段映射后的目标表副本
        mapped_target_df = target_df.copy()
        reverse_mapping = {v: k for k, v in field_mapping.items()}
        
        # 重命名目标表字段以匹配源表
        for target_field, source_field in reverse_mapping.items():
            if target_field in mapped_target_df.columns:
                mapped_target_df = mapped_target_df.rename(columns={target_field: source_field})
        
        # 使用映射后的字段创建索引
        source_index = source_df.set_index(key_fields)
        target_index = mapped_target_df.set_index(key_fields)
    else:
        # 如果没有字段映射，直接使用原始字段
        source_index = source_df.set_index(key_fields)
        target_index = target_df.set_index(key_fields)
    
    # 检查数据丢失 (源数据中有但目标数据中没有的记录)
    source_keys = set(source_index.index)
    target_keys = set(target_index.index)
    
    data_loss_keys = source_keys - target_keys
    if data_loss_keys:
        for key in data_loss_keys:
            if isinstance(key, tuple):
                key_dict = dict(zip(key_fields, key))
            else:
                key_dict = {key_fields[0]: key}
            
            # 获取源数据记录
            source_record = source_index.loc[key]
            if isinstance(source_record, pd.Series):
                source_record = source_record.to_dict()
            
            result['data_loss'].append({
                'key': key_dict,
                'source_data': source_record,
                'reason': 'Record exists in source but not in target'
            })
    
    # 检查值差异 (两个表中都存在但值不匹配的记录)
    common_keys = source_keys & target_keys
    
    for key in common_keys:
        source_record = source_index.loc[key]
        target_record = target_index.loc[key]
        
        if isinstance(source_record, pd.Series):
            source_record = source_record.to_dict()
        if isinstance(target_record, pd.Series):
            target_record = target_record.to_dict()
        
        differences = {}
        has_diff = False
        
        # 比较映射字段的值
        if field_mapping:
            for source_field, target_field in field_mapping.items():
                if source_field in source_record and source_field in target_record:
                    source_value = source_record[source_field]
                    target_value = target_record[source_field]  # 使用映射后的字段名
                    
                    # 处理NaN值
                    if pd.isna(source_value) and pd.isna(target_value):
                        continue
                    elif pd.isna(source_value) or pd.isna(target_value):
                        differences[source_field] = {
                            'source_value': source_value,
                            'target_value': target_value,
                            'target_field': target_field
                        }
                        has_diff = True
                    elif str(source_value) != str(target_value):
                        differences[source_field] = {
                            'source_value': source_value,
                            'target_value': target_value,
                            'target_field': target_field
                        }
                        has_diff = True
        else:
            # 如果没有字段映射，比较共同字段
            common_fields = set(source_record.keys()) & set(target_record.keys())
            for field in common_fields:
                if field not in key_fields:  # 跳过关键字段
                    source_value = source_record[field]
                    target_value = target_record[field]
                    
                    # 处理NaN值
                    if pd.isna(source_value) and pd.isna(target_value):
                        continue
                    elif pd.isna(source_value) or pd.isna(target_value):
                        differences[field] = {
                            'source_value': source_value,
                            'target_value': target_value,
                            'target_field': field
                        }
                        has_diff = True
                    elif str(source_value) != str(target_value):
                        differences[field] = {
                            'source_value': source_value,
                            'target_value': target_value,
                            'target_field': field
                        }
                        has_diff = True
        
        if has_diff:
            if isinstance(key, tuple):
                key_dict = dict(zip(key_fields, key))
            else:
                key_dict = {key_fields[0]: key}
            
            result['value_diff'].append({
                'key': key_dict,
                'source_data': source_record,
                'target_data': target_record,
                'differences': differences
            })
    
    # 生成摘要信息
    result['summary'] = {
        'source_total_records': len(source_df),
        'target_total_records': len(target_df),
        'data_loss_count': len(result['data_loss']),
        'value_diff_count': len(result['value_diff']),
        'matching_records': len(common_keys) - len(result['value_diff']),
        'field_mapping': field_mapping,
        'key_fields': key_fields
    }
    
    logger.info(f"Comparison completed: {result['summary']}")
    
    return result

def generate_excel_report(comparison_result: Dict[str, Any]) -> str:
    """
    生成Excel报告
    
    Args:
        comparison_result: 比较结果
        
    Returns:
        Excel文件路径
    """
    try:
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            # 创建数据丢失工作表
            if comparison_result['data_loss']:
                data_loss_df = create_data_loss_dataframe(comparison_result['data_loss'])
                data_loss_df.to_excel(writer, sheet_name='Data_Loss', index=False)
                apply_data_loss_styling(writer, data_loss_df)
            
            # 创建值差异工作表
            if comparison_result['value_diff']:
                value_diff_df = create_value_diff_dataframe(comparison_result['value_diff'])
                value_diff_df.to_excel(writer, sheet_name='Value_Differences', index=False)
                apply_value_diff_styling(writer, value_diff_df)
            
            # 创建摘要工作表
            summary_df = create_summary_dataframe(comparison_result['summary'])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            apply_summary_styling(writer, summary_df)
        
        logger.info(f"Excel report generated: {temp_file.name}")
        return temp_file.name
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise

def create_data_loss_dataframe(data_loss: List[Dict]) -> pd.DataFrame:
    """创建数据丢失DataFrame"""
    if not data_loss:
        return pd.DataFrame()
    
    rows = []
    for item in data_loss:
        row = {}
        # 添加关键字段
        for field, value in item['key'].items():
            row[f'Key_{field}'] = value
        
        # 添加源数据字段
        for field, value in item['source_data'].items():
            row[f'Source_{field}'] = value
        
        row['Reason'] = item['reason']
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_value_diff_dataframe(value_diff: List[Dict]) -> pd.DataFrame:
    """创建值差异DataFrame"""
    if not value_diff:
        return pd.DataFrame()
    
    rows = []
    for item in value_diff:
        row = {}
        # 添加关键字段
        for field, value in item['key'].items():
            row[f'Key_{field}'] = value
        
        # 添加源数据字段
        for field, value in item['source_data'].items():
            row[f'Source_{field}'] = value
        
        # 添加目标数据字段
        for field, value in item['target_data'].items():
            row[f'Target_{field}'] = value
        
        # 添加差异信息
        for field, diff_info in item['differences'].items():
            row[f'Diff_{field}_Source'] = diff_info['source_value']
            row[f'Diff_{field}_Target'] = diff_info['target_value']
            row[f'Diff_{field}_TargetField'] = diff_info['target_field']
        
        rows.append(row)
    
    return pd.DataFrame(rows)

def create_summary_dataframe(summary: Dict) -> pd.DataFrame:
    """创建摘要DataFrame"""
    summary_data = []
    for key, value in summary.items():
        if key not in ['field_mapping', 'key_fields']:
            summary_data.append({'Metric': key, 'Value': value})
    
    # 添加字段映射信息
    for source_field, target_field in summary.get('field_mapping', {}).items():
        summary_data.append({
            'Metric': f'Field_Mapping_{source_field}',
            'Value': target_field
        })
    
    # 添加关键字段信息
    for field in summary.get('key_fields', []):
        summary_data.append({
            'Metric': f'Key_Field_{field}',
            'Value': 'Yes'
        })
    
    return pd.DataFrame(summary_data) 

def apply_data_loss_styling(writer, df):
    """为数据丢失工作表应用样式"""
    if df.empty:
        return
        
    worksheet = writer.sheets['Data_Loss']
    workbook = writer.book
    
    # 定义样式 - 使用openpyxl.styles直接创建样式对象
    header_style = openpyxl.styles.NamedStyle(name='HeaderStyle')
    header_style.font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=12)
    header_style.fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    header_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 数据行样式
    data_style = openpyxl.styles.NamedStyle(name='DataStyle')
    data_style.fill = openpyxl.styles.PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
    data_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 应用表头样式
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.style = header_style
    
    # 应用数据行样式
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.style = data_style
    
    # 调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_value_diff_styling(writer, df):
    """为值差异工作表应用样式"""
    if df.empty:
        return
        
    worksheet = writer.sheets['Value_Differences']
    workbook = writer.book
    
    # 定义样式 - 使用openpyxl.styles直接创建样式对象
    header_style = openpyxl.styles.NamedStyle(name='ValueHeaderStyle')
    header_style.font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=12)
    header_style.fill = openpyxl.styles.PatternFill(start_color='C5504B', end_color='C5504B', fill_type='solid')
    header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    header_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 关键字段样式
    key_style = openpyxl.styles.NamedStyle(name='KeyStyle')
    key_style.fill = openpyxl.styles.PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    key_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 源数据样式
    source_style = openpyxl.styles.NamedStyle(name='SourceStyle')
    source_style.fill = openpyxl.styles.PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    source_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 目标数据样式
    target_style = openpyxl.styles.NamedStyle(name='TargetStyle')
    target_style.fill = openpyxl.styles.PatternFill(start_color='E1D5E7', end_color='E1D5E7', fill_type='solid')
    target_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 差异值样式 - 突出显示
    diff_style = openpyxl.styles.NamedStyle(name='DiffStyle')
    diff_style.fill = openpyxl.styles.PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    diff_style.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    diff_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thick', color='FF0000'),
        right=openpyxl.styles.Side(style='thick', color='FF0000'),
        top=openpyxl.styles.Side(style='thick', color='FF0000'),
        bottom=openpyxl.styles.Side(style='thick', color='FF0000')
    )
    
    # 应用表头样式
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.style = header_style
    
    # 应用数据行样式
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            column_name = df.columns[col - 1]
            
            # 根据列名应用不同样式
            if column_name.startswith('Key_'):
                cell.style = key_style
            elif column_name.startswith('Source_'):
                cell.style = source_style
            elif column_name.startswith('Target_'):
                cell.style = target_style
            elif column_name.startswith('Diff_'):
                cell.style = diff_style
            else:
                # 默认样式
                default_style = openpyxl.styles.NamedStyle(name='DefaultStyle')
                default_style.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                cell.style = default_style
    
    # 调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_summary_styling(writer, df):
    """为摘要工作表应用样式"""
    if df.empty:
        return
        
    worksheet = writer.sheets['Summary']
    workbook = writer.book
    
    # 定义样式 - 使用openpyxl.styles直接创建样式对象
    header_style = openpyxl.styles.NamedStyle(name='SummaryHeaderStyle')
    header_style.font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=12)
    header_style.fill = openpyxl.styles.PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    header_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 数据行样式
    data_style = openpyxl.styles.NamedStyle(name='SummaryDataStyle')
    data_style.fill = openpyxl.styles.PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    data_style.border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 应用表头样式
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.style = header_style
    
    # 应用数据行样式
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.style = data_style
    
    # 调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width